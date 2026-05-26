"""
migrate_contacts.py
===================
Migrates contacts from the legacy Business_Cards_Catalogue.xlsx into the
NetworkTracker.xlsx format (single "Contacts" table on the Contacts sheet).

Handles:
  - Standard person-tabs (Ravi, Philip, Davin, etc.) with 8-column layout
  - "Sheet2" (phone-contacts export) with a different 42-column layout
  - "Others" tab whose 8th column is SOURCE instead of DATE ADDED
  - Empty tabs (Eason, Jordan) — skipped automatically
  - Deletion of the 3 sample/demo rows already in the target file

Column order in the output (matches NetworkTracker schema):
  id | name | company | position | email | phoneMobile | phoneOffice |
  linkedin | type | heat | frequency | eventMet | notes | owners |
  dateAdded | lastTouched
"""

import copy
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

LEGACY_PATH = "/root/.claude/uploads/db775b41-f0dd-439b-b7dc-5d890024d394/84d44788-Business_Cards_Catalogue.xlsx"
TARGET_PATH = "/root/.claude/uploads/db775b41-f0dd-439b-b7dc-5d890024d394/7ade580f-NetworkTracker.xlsx"
OUTPUT_PATH = "/home/user/sage3-network-tracker/NetworkTracker.xlsx"

HEADERS = [
    "id", "name", "company", "position", "email", "phoneMobile",
    "phoneOffice", "linkedin", "type", "heat", "frequency", "eventMet",
    "notes", "owners", "dateAdded", "lastTouched",
]

STANDARD_LEGACY_COLS = ["NO", "NAME", "COMPANY", "POSITION", "EMAIL",
                        "PHONE (MOBILE)", "PHONE (OFFICE)"]

SHEET2_COL_MAP = {
    "Full Name": "name",
    "Company1": "company",
    "Job Title1": "position",
    "Email1": "email",
    "Mobile1": "phoneMobile",
    "Tel1": "phoneOffice",
    "Date Created": "dateAdded",
}


def normalise_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else ""


def normalise_str(val):
    if val is None:
        return ""
    return str(val).strip()


def read_standard_tab(ws, sheet_name):
    rows_out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name_val = row[1]
        if name_val is None or str(name_val).strip() == "":
            continue

        is_others = (sheet_name == "Others")
        date_added = "" if is_others else normalise_date(row[7] if len(row) > 7 else None)

        rows_out.append({
            "name": normalise_str(row[1]),
            "company": normalise_str(row[2]),
            "position": normalise_str(row[3]),
            "email": normalise_str(row[4]),
            "phoneMobile": normalise_str(row[5]),
            "phoneOffice": normalise_str(row[6]),
            "dateAdded": date_added,
            "owners": sheet_name,
        })
    return rows_out


def read_sheet2(ws):
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {}
    for src_col, tgt_col in SHEET2_COL_MAP.items():
        try:
            col_idx[tgt_col] = header.index(src_col)
        except ValueError:
            col_idx[tgt_col] = None

    rows_out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name_idx = col_idx.get("name")
        if name_idx is None or row[name_idx] is None or str(row[name_idx]).strip() == "":
            continue

        entry = {"owners": "Sheet2"}
        for tgt_col, idx in col_idx.items():
            if idx is not None and idx < len(row):
                if tgt_col == "dateAdded":
                    entry[tgt_col] = normalise_date(row[idx])
                else:
                    entry[tgt_col] = normalise_str(row[idx])
            else:
                entry[tgt_col] = ""

        for field in ["name", "company", "position", "email", "phoneMobile", "phoneOffice", "dateAdded"]:
            entry.setdefault(field, "")

        rows_out.append(entry)
    return rows_out


def build_output_row(seq_id, entry):
    date_added = entry.get("dateAdded", "")
    return [
        str(seq_id).zfill(3),            # id
        entry.get("name", ""),            # name
        entry.get("company", ""),         # company
        entry.get("position", ""),        # position
        entry.get("email", ""),           # email
        entry.get("phoneMobile", ""),     # phoneMobile
        entry.get("phoneOffice", ""),     # phoneOffice
        "",                               # linkedin
        "",                               # type
        "",                               # heat
        "biannual",                       # frequency
        "",                               # eventMet
        "",                               # notes
        entry.get("owners", ""),          # owners
        date_added,                       # dateAdded
        date_added,                       # lastTouched = dateAdded
    ]


def main():
    legacy_wb = openpyxl.load_workbook(LEGACY_PATH)
    target_wb = openpyxl.load_workbook(TARGET_PATH)

    all_contacts = []
    tab_summary = {}

    for sheet_name in legacy_wb.sheetnames:
        ws = legacy_wb[sheet_name]

        if sheet_name == "Sheet2":
            rows = read_sheet2(ws)
        else:
            rows = read_standard_tab(ws, sheet_name)

        tab_summary[sheet_name] = len(rows)
        all_contacts.extend(rows)

    legacy_wb.close()

    # --- Write to target workbook ---
    ws_contacts = target_wb["Contacts"]

    # Remove existing table definition (we'll recreate it)
    for tbl_name in list(ws_contacts.tables.keys()):
        del ws_contacts.tables[tbl_name]

    # Delete existing data rows (keep header in row 1)
    if ws_contacts.max_row > 1:
        ws_contacts.delete_rows(2, ws_contacts.max_row - 1)

    # Preserve header formatting from row 1
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Ensure headers are correct
    for col_idx, header_val in enumerate(HEADERS, start=1):
        cell = ws_contacts.cell(row=1, column=col_idx, value=header_val)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Write all contact rows
    data_font = Font(name="Calibri", size=11)
    for i, entry in enumerate(all_contacts, start=1):
        row_data = build_output_row(i, entry)
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_contacts.cell(row=i + 1, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    total_rows = len(all_contacts)

    # Set column widths for readability
    col_widths = {
        "A": 8, "B": 30, "C": 40, "D": 35, "E": 35,
        "F": 20, "G": 20, "H": 25, "I": 16, "J": 10,
        "K": 14, "L": 15, "M": 30, "N": 18, "O": 14, "P": 14,
    }
    for col_letter, width in col_widths.items():
        ws_contacts.column_dimensions[col_letter].width = width

    # Recreate the formal Excel Table
    table_ref = f"A1:P{total_rows + 1}"
    table = Table(displayName="Contacts", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws_contacts.add_table(table)

    # Add data validation for 'type' (column I)
    dv_type = DataValidation(
        type="list",
        formula1='"client,capital_provider,partner"',
        allow_blank=True,
    )
    dv_type.error = "Please select client, capital_provider, or partner"
    dv_type.errorTitle = "Invalid Type"
    dv_type.add(f"I2:I{total_rows + 1}")
    ws_contacts.add_data_validation(dv_type)

    # Add data validation for 'heat' (column J)
    dv_heat = DataValidation(
        type="list",
        formula1='"hot,warm,cold"',
        allow_blank=True,
    )
    dv_heat.error = "Please select hot, warm, or cold"
    dv_heat.errorTitle = "Invalid Heat"
    dv_heat.add(f"J2:J{total_rows + 1}")
    ws_contacts.add_data_validation(dv_heat)

    # Add data validation for 'frequency' (column K)
    dv_freq = DataValidation(
        type="list",
        formula1='"biannual,quarterly,monthly,asneeded"',
        allow_blank=True,
    )
    dv_freq.add(f"K2:K{total_rows + 1}")
    ws_contacts.add_data_validation(dv_freq)

    # Clean up Interactions sample data — keep one blank row so the table is valid
    ws_interactions = target_wb["Interactions"]
    for tbl_name in list(ws_interactions.tables.keys()):
        tbl = ws_interactions.tables[tbl_name]
        old_style = tbl.tableStyleInfo
        del ws_interactions.tables[tbl_name]
    if ws_interactions.max_row > 1:
        ws_interactions.delete_rows(2, ws_interactions.max_row - 1)
    for col in range(1, 8):
        ws_interactions.cell(row=2, column=col, value="")
    new_tbl = Table(displayName="Interactions", ref="A1:G2")
    new_tbl.tableStyleInfo = old_style
    ws_interactions.add_table(new_tbl)

    target_wb.save(OUTPUT_PATH)
    target_wb.close()

    # --- Summary ---
    print("=" * 60)
    print("MIGRATION COMPLETE")
    print("=" * 60)
    print(f"Output: {OUTPUT_PATH}")
    print(f"Total contacts migrated: {total_rows}")
    print()
    print("Breakdown by tab:")
    for tab, count in tab_summary.items():
        status = "EMPTY" if count == 0 else f"{count} contacts"
        print(f"  {tab:20s} → {status}")
    print()
    print("Target columns: " + ", ".join(HEADERS))
    print("Table name: Contacts | Style: TableStyleMedium9")
    print("Sample rows deleted: YES (3 demo rows removed)")
    print("Data validations added: type, heat, frequency")


if __name__ == "__main__":
    main()
